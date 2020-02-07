import openpyxl
import pandas as pd

###READ POSTAL CODES###

book = openpyxl.load_workbook('../references/infos_membres.xlsx')
sheet = book["Feuil1"]

mylist = []
for col in sheet['C']:
     mylist.append(col.value)

###GET CIRCONSCRIPTIONS###

circonscriptions = []

for i in range(2):
    #WRITE IN THE DOC
    book2 = openpyxl.load_workbook('../references/CP_CIRC_MUN_MRC_RA.xlsm')
    sheet2 = book2["Interrogation"]
    cell = sheet2.cell(row=3, column=2)
    cell.value = mylist[i]
    book2.save('../references/CP_CIRC_MUN_MRC_RA.xlsm')
    book2.close()
    
    #REOPEN TO GET RESULTED VALUE
    book3 = openpyxl.load_workbook('../references/CP_CIRC_MUN_MRC_RA.xlsm', data_only=True)
    sheet3 = book3["Interrogation"]
    cell_answer = sheet3.cell(row=6, column=3)
    circonscriptions.append(cell_answer.internal_value)
    book3.close()

for c in circonscriptions:
    print(c)
    
 #:RE-OPEN TO GET MODIFIED INFO
   #:data =pd.read_excel('./CP_CIRC_MUN_MRC_RA.xlsm', sheet_name='Interrogation', index_col=None, usecols = "C", skiprows =5, nrows=1)
   # print(data) #book2 = openpyxl.load_workbook('../references/CP_CIRC_MUN_MRC_RA.xlsm', data_only=True)
     


